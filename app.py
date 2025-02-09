from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from io import BytesIO
import os

app = Flask(__name__)
app.secret_key = '12345'  # Cambia este valor por uno seguro

# Configurar la base de datos SQLite
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///trabajadores.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# Definir el modelo de datos
class Trabajador(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), unique=True, nullable=False)
    procesos = db.Column(db.Integer, default=0)

# Crear las tablas (si no existen)
with app.app_context():
    db.create_all()

# Ruta principal: muestra el dashboard, el formulario y el listado
@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        try:
            procesos = int(request.form.get('procesos', 0))
        except ValueError:
            flash("El valor de procesos debe ser numérico", "error")
            return redirect(url_for('index'))
        if not nombre:
            flash("El nombre no puede estar vacío", "error")
            return redirect(url_for('index'))
        # Buscar si ya existe un trabajador (sin distinguir mayúsculas/minúsculas)
        trabajador = Trabajador.query.filter(
            db.func.lower(Trabajador.nombre) == nombre.lower()
        ).first()
        if trabajador:
            trabajador.procesos = procesos
            flash(f"Se actualizó {nombre}", "info")
        else:
            trabajador = Trabajador(nombre=nombre, procesos=procesos)
            db.session.add(trabajador)
            flash(f"Se agregó {nombre}", "info")
        db.session.commit()
        return redirect(url_for('index'))
    trabajadores = Trabajador.query.all()
    return render_template('index.html', persons=trabajadores)

# Ruta para editar un trabajador (usando su ID)
@app.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    trabajador = Trabajador.query.get_or_404(id)
    if request.method == 'POST':
        nuevo_nombre = request.form.get('nombre', '').strip()
        try:
            procesos = int(request.form.get('procesos', 0))
        except ValueError:
            flash("El valor de procesos debe ser numérico", "error")
            return redirect(url_for('editar', id=id))
        if not nuevo_nombre:
            flash("El nombre no puede estar vacío", "error")
            return redirect(url_for('editar', id=id))
        # Validar que no se duplique el nombre (ignorando mayúsculas)
        if nuevo_nombre.lower() != trabajador.nombre.lower():
            existente = Trabajador.query.filter(
                db.func.lower(Trabajador.nombre) == nuevo_nombre.lower()
            ).first()
            if existente and existente.id != id:
                flash("Ya existe otro trabajador con ese nombre.", "error")
                return redirect(url_for('editar', id=id))
        trabajador.nombre = nuevo_nombre
        trabajador.procesos = procesos
        db.session.commit()
        flash("Datos actualizados", "info")
        return redirect(url_for('index'))
    return render_template('editar.html', person=trabajador)

# Endpoint para actualizar (incrementar o decrementar) el contador de procesos vía AJAX
@app.route('/actualizar/<int:id>', methods=['POST'])
def actualizar(id):
    accion = request.args.get('accion')
    trabajador = Trabajador.query.get_or_404(id)
    if accion == 'mas':
        trabajador.procesos += 1
    elif accion == 'menos' and trabajador.procesos > 0:
        trabajador.procesos -= 1
    db.session.commit()
    return jsonify({'procesos': trabajador.procesos})

# Ruta para importar datos desde Excel
@app.route('/importar', methods=['GET', 'POST'])
def importar():
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file:
            flash("No se seleccionó ningún archivo", "error")
            return redirect(url_for('importar'))
        try:
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
        except Exception as e:
            flash(f"Error al leer el archivo: {e}", "error")
            return redirect(url_for('importar'))
        updated_count = 0
        # Se asume que la primera fila tiene encabezados: Nombre, Tipo, Cantidad
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None or row[1] is None:
                continue
            nombre = str(row[0]).strip()
            tipo = str(row[1]).strip().capitalize()  # Se espera "Procesos"
            if tipo != "Procesos":
                continue
            quantity = 1
            if len(row) >= 3 and row[2] is not None:
                try:
                    quantity = int(row[2])
                except ValueError:
                    quantity = 1
            trabajador = Trabajador.query.filter(
                db.func.lower(Trabajador.nombre) == nombre.lower()
            ).first()
            if trabajador:
                trabajador.procesos = quantity
            else:
                trabajador = Trabajador(nombre=nombre, procesos=quantity)
                db.session.add(trabajador)
            updated_count += 1
        db.session.commit()
        flash(f"Se importaron {updated_count} registros desde Excel", "info")
        return redirect(url_for('index'))
    return render_template('importar.html')

# Ruta para exportar datos a Excel
@app.route('/exportar')
def exportar():
    from openpyxl import Workbook
    wb = Workbook()
    sheet = wb.active
    # Escribir encabezados
    sheet.append(["Nombre", "Procesos"])
    trabajadores = Trabajador.query.all()
    for t in trabajadores:
        sheet.append([t.nombre, t.procesos])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="exportado.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/eliminar/<int:id>', methods=['POST'])
def eliminar(id):
    trabajador = Trabajador.query.get_or_404(id)
    db.session.delete(trabajador)
    db.session.commit()
    return jsonify({'success': True})

# Ruta para obtener datos en formato JSON (para la gráfica)
@app.route('/data')
def data():
    trabajadores = Trabajador.query.all()
    data = {
        "nombres": [t.nombre for t in trabajadores],
        "procesos": [t.procesos for t in trabajadores]
    }
    return jsonify(data)

if __name__ == '__main__':
    # Para permitir el acceso externo, se utiliza host='0.0.0.0'
    app.run(host='0.0.0.0', port=5000, debug=True)
